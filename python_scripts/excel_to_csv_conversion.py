import os
import csv
from openpyxl import load_workbook

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
INPUT_DIR = os.path.join(ROOT_DIR, "source_data")   # sumup input .xlsx files here
OUTPUT_DIR = os.path.join(ROOT_DIR, "csv_raw_files")

def convert_xlsx_to_csv(xlsx_file_path, csv_file_path):
    workbook = load_workbook(filename=xlsx_file_path, read_only=True, data_only=True)
    sheet = workbook.active

    with open(csv_file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(row)

    print(f"✅ Converted {os.path.basename(xlsx_file_path)} → {os.path.basename(csv_file_path)}")

def convert_all_excels(input_dir=INPUT_DIR, output_dir=OUTPUT_DIR):
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    for file_name in os.listdir(input_dir):
        if file_name.endswith(".xlsx"):
            xlsx_path = os.path.join(input_dir, file_name)
            csv_name = file_name.replace(".xlsx", ".csv")
            csv_path = os.path.join(output_dir, csv_name)
            convert_xlsx_to_csv(xlsx_path, csv_path)

if __name__ == "__main__":
    convert_all_excels()
