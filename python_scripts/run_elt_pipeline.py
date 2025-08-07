import os
import subprocess
import glob
import openpyxl
from openpyxl import load_workbook
import csv


ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
INPUT_DIR = os.path.join(ROOT_DIR, "source_data")   # sumup input .xlsx files here
SEED_DIR = os.path.join(ROOT_DIR, "seeds")

os.makedirs(SEED_DIR, exist_ok=True)

def convert_xlsx_to_csv(xlsx_path, SEED_DIR):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    csv_path = os.path.join(SEED_DIR, f"{base_name}.csv")

    with open(csv_path, 'w', newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, quoting=csv.QUOTE_MINIMAL)
        for row in sheet.iter_rows(values_only=True):
            clean_row = [str(cell).strip() if cell is not None else "" for cell in row]
            writer.writerow(clean_row)

    print(f"✅ Converted: {xlsx_path} → {csv_path}")
    return csv_path

def convert_all_xlsx_in_folder(INPUT_DIR, SEED_DIR):
    for file in os.listdir(INPUT_DIR):
        if file.endswith(".xlsx"):
            xlsx_path = os.path.join(INPUT_DIR, file)
            convert_xlsx_to_csv(xlsx_path, SEED_DIR)

def run_dbt():
    print("🚀 Running dbt pipeline...")
    subprocess.run(["dbt", "seed", "--full-refresh"], check=True, cwd=ROOT_DIR)
    subprocess.run(["dbt", "build"], check=True, cwd=ROOT_DIR)
    subprocess.run(["dbt", "run"], check=True, cwd=ROOT_DIR)
    print("✅ DBT pipeline completed successfully.")

def main():
    print("🔧 Starting data pipeline...")
    convert_all_xlsx_in_folder(INPUT_DIR, SEED_DIR)
    run_dbt()
    print("🎉 All done!")   

if __name__ == "__main__":
    main()